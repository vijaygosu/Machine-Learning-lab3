import numpy as np
A=np.array([[20,6,2],[16,3,6],[27,6,2],[19,1,2],[24,4,2],[22,1,5],[15,4,2],[18,4,2],[21,1,4],[16,2,4]])
C=np.array([[386],[289],[393],[110],[280],[167],[271],[274],[148],[192]])
print(A,C)
print(A.shape,C.shape)  
import openpyxl

file_path = "C:\\Users\\vijay\\Documents\\sem4\\ML\\Lab Session1 Data.xlsx"
sheet_name = "Purchase data"

wb = openpyxl.load_workbook(file_path)
sheet = wb[sheet_name]
start_row = 2 
start_col = 2
end_col = sheet.max_column

data = []

for row_idx in range(start_row, sheet.max_row + 1):
    row_data = []
    for col_idx in range(start_col, end_col + 1):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        row_data.append(cell_value)
    data.append(row_data)

print("Extracted data:")
print(data)

